import sys

from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QComboBox, QPushButton, QVBoxLayout, QWidget, \
    QMessageBox, QGroupBox, QFormLayout
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook, load_workbook
import webbrowser


def create_patient_folder(patient_name):
    folder_name = patient_name.replace(" ", "_")  # Replace spaces with underscores
    os.makedirs(folder_name, exist_ok=True)
    return folder_name


def load_excel_file(file_path):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()
        workbook.save(file_path)

    return workbook


def check_existing_mobile_number(mobile_number):
    for folder_name in os.listdir('.'):
        if os.path.isdir(folder_name):
            file_names = [file for file in os.listdir(folder_name) if file.endswith('.xlsx')]
            for file_name in file_names:
                file_path = os.path.join(folder_name, file_name)
                workbook = load_workbook(file_path)
                sheet = workbook.active
                if sheet.cell(row=2, column=4).value == mobile_number:
                    return folder_name

    return None


def add_patient():
    name = line_edit_name.text()
    age = line_edit_age.text()
    gender = combo_box_gender.currentText()
    mobile_number = line_edit_mobile.text()
    disease = line_edit_disease.text()
    town = line_edit_town.text()
    fee = line_edit_fee.text()

    if name and age and gender and mobile_number and disease and town and fee:
        if len(mobile_number) != 10:
            QMessageBox.warning(window, "Error", "Mobile number should be 10 digits.")
            return

        try:
            fee = float(fee)
        except ValueError:
            QMessageBox.warning(window, "Error", "Fee should be a valid floating-point number.")
            return

        existing_folder = check_existing_mobile_number(mobile_number)
        if existing_folder:
            folder_name = existing_folder
            current_date = datetime.now().strftime('%Y%m%d_%H%M%S')  # Add current time to the filename
        else:
            folder_name = create_patient_folder(name)
            current_date = datetime.now().strftime('%Y%m%d')

        file_name = os.path.join(folder_name, f"{name.replace(' ', '_')}_{current_date}.xlsx")

        workbook = load_excel_file(file_name)
        sheet = workbook.active

        if sheet.max_row == 1:
            sheet.append(["Name", "Age", "Gender", "Mobile Number", "Disease", "Town", "Fee"])  # Add header row

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = name
        sheet.cell(row=next_row, column=2).value = age
        sheet.cell(row=next_row, column=3).value = gender
        sheet.cell(row=next_row, column=4).value = mobile_number
        sheet.cell(row=next_row, column=5).value = disease
        sheet.cell(row=next_row, column=6).value = town
        sheet.cell(row=next_row, column=7).value = fee

        workbook.save(file_name)

        QMessageBox.information(window, "Success", "Patient added successfully.")
        line_edit_name.clear()
        line_edit_age.clear()
        combo_box_gender.setCurrentIndex(0)
        line_edit_mobile.clear()
        line_edit_disease.clear()
        line_edit_town.clear()
        line_edit_fee.clear()
    else:
        QMessageBox.warning(window, "Error", "Please fill in all the fields.")


def search_patient():
    mobile_number = line_edit_search_mobile.text()
    if mobile_number:
        existing_folder = check_existing_mobile_number(mobile_number)
        if existing_folder:
            folder_path = os.path.abspath(existing_folder)
            webbrowser.open(folder_path)
        else:
            QMessageBox.warning(window, "Patient Not Found", "No patient found with the given mobile number.")
    else:
        QMessageBox.warning(window, "Error", "Please enter a mobile number to search.")


app = QApplication(sys.argv)
icon_path = "C:/Users/Fatma/OneDrive/Documents/GitHub/pythonProject2/logo.ico"
# Replace with the actual path to your icon file
app.setWindowIcon(QIcon(icon_path))
window = QMainWindow()
window.setWindowTitle("Patient Management System")

central_widget = QWidget(window)
window.setCentralWidget(central_widget)

layout = QVBoxLayout()
central_widget.setLayout(layout)

group_box_patient_info = QGroupBox("Patient Information")
form_layout = QFormLayout()
group_box_patient_info.setLayout(form_layout)

line_edit_name = QLineEdit()
line_edit_age = QLineEdit()
combo_box_gender = QComboBox()
combo_box_gender.addItems(["Male", "Female", "Other"])
line_edit_mobile = QLineEdit()
line_edit_disease = QLineEdit()
line_edit_town = QLineEdit()
line_edit_fee = QLineEdit()

form_layout.addRow(QLabel("Name:"), line_edit_name)
form_layout.addRow(QLabel("Age:"), line_edit_age)
form_layout.addRow(QLabel("Gender:"), combo_box_gender)
form_layout.addRow(QLabel("Mobile Number:"), line_edit_mobile)
form_layout.addRow(QLabel("Disease:"), line_edit_disease)
form_layout.addRow(QLabel("Town:"), line_edit_town)
form_layout.addRow(QLabel("Fee:"), line_edit_fee)

button_add = QPushButton("Add Patient")
button_add.clicked.connect(add_patient)

group_box_patient_info.layout().addWidget(button_add)
layout.addWidget(group_box_patient_info)

group_box_search_patient = QGroupBox("Search Patient")
form_layout_search = QFormLayout()
group_box_search_patient.setLayout(form_layout_search)

line_edit_search_mobile = QLineEdit()
button_search = QPushButton("Search")
button_search.clicked.connect(search_patient)

form_layout_search.addRow(QLabel("Search Patient by Mobile Number:"), line_edit_search_mobile)
form_layout_search.addRow(button_search)

layout.addWidget(group_box_search_patient)

window.show()
sys.exit(app.exec_())
